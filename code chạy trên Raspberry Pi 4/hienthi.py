# -*- coding: utf-8 -*-
import os, sys, cv2, numpy as np, shutil
from PyQt5 import QtWidgets, uic, QtCore, QtGui
from PyQt5.QtCore import QProcess, QTimer, Qt, QDateTime, pyqtSignal, QObject, QThread
from PyQt5.QtGui import QImage, QPixmap, QFont
from picamera2 import Picamera2
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

os.environ["QT_QPA_PLATFORM"] = "xcb"

class ExportWorker(QObject):
    finished = pyqtSignal(str)
    def __init__(self, excel_file):
        super().__init__()
        self.excel_file = excel_file

    def run(self):
        results = []
        try:
            res = os.system(f"rclone copy {self.excel_file} gdrive:DoAn_Lidar/")
            results.append("✅ Google Drive: OK") if res == 0 else results.append("❌ Google Drive: Lỗi Rclone")
        except: results.append("❌ Google Drive: Lỗi kết nối")

        try:
            username = os.getlogin()
            usb_base = f"/media/{username}/"
            if os.path.exists(usb_base) and os.listdir(usb_base):
                usb_path = os.path.join(usb_base, os.listdir(usb_base)[0])
                dest = os.path.join(usb_path, "BaoCao_KetQua.xlsx")
                shutil.copy(self.excel_file, dest)
                results.append(f"✅ USB ({os.listdir(usb_base)[0]}): OK")
            else: results.append("❌ USB: Không tìm thấy")
        except: results.append("❌ USB: Lỗi sao chép")

        self.finished.emit("\n".join(results))

class CameraWorker(QThread):
    frame_ready = pyqtSignal(np.ndarray, list)
    def __init__(self, picam2, net, classes):
        super().__init__()
        self.picam2 = picam2
        self.net = net
        self.CLASSES = classes
        self.running = True

    def run(self):
        while self.running:
            try:
                frame = self.picam2.capture_array()
                if frame is None: continue
                
                frame_disp = cv2.resize(frame, (580, 330))
                h, w = frame_disp.shape[:2]
                found = []
                
                if self.net is not None:
                    blob = cv2.dnn.blobFromImage(frame_disp, 0.007843, (300, 300), 127.5)
                    self.net.setInput(blob)
                    detections = self.net.forward()
                    for i in range(detections.shape[2]):
                        if detections[0,0,i,2] > 0.6:
                            idx = int(detections[0,0,i,1])
                            name = self.CLASSES[idx].upper()
                            box = detections[0,0,i,3:7] * np.array([w,h,w,h])
                            (sX, sY, eX, eY) = box.astype("int")
                            cv2.rectangle(frame_disp, (sX, sY), (eX, eY), (0, 255, 0), 2)
                            cv2.putText(frame_disp, name, (sX, sY-10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                            found.append(name)
                
                self.frame_ready.emit(frame_disp, found)
            except: pass
            self.msleep(10)

    def stop(self):
        self.running = False
        self.wait()

class LidarApp(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("hienthi.ui", self)
        
        self.BASE_DIR = "/home/doan2026/Desktop/qt"
        self.EXCEL_FILE = os.path.join(self.BASE_DIR, "data_log.xlsx")
        self.lidar_full_data = [0.0] * 360
        self.last_val = 0 
        self.latest_frame_save = None  
        
        self.CLASSES = ["background", "aeroplane", "bicycle", "bird", "boat",
                        "bottle", "bus", "car", "cat", "chair", "cow", "diningtable",
                        "dog", "horse", "motorbike", "person", "pottedplant", "sheep",
                        "sofa", "train", "tvmonitor"]

        self.init_hardware()
        
        self.pushButton.clicked.connect(self.start_lidar)       
        self.pushButton_2.clicked.connect(self.stop_all)      
        self.save_btn.clicked.connect(self.save_data_to_excel) 
        self.export_btn.clicked.connect(self.start_export_thread) 

        self.process = QProcess(self)
        self.process.readyReadStandardOutput.connect(self.read_lidar)
        
        self.cam_scene = QtWidgets.QGraphicsScene(self)
        self.CameraView.setScene(self.cam_scene)

        # --- CÀI ĐẶT PHÍM TẮT ĐIỀU KHIỂN BẢN ĐỒ ---
        # Phím mũi tên Lên và phím + để Zoom In
        QtWidgets.QShortcut(QtGui.QKeySequence("Up"), self).activated.connect(self.zoom_in)
        QtWidgets.QShortcut(QtGui.QKeySequence("+"), self).activated.connect(self.zoom_in)
        
        # Phím mũi tên Xuống và phím - để Zoom Out
        QtWidgets.QShortcut(QtGui.QKeySequence("Down"), self).activated.connect(self.zoom_out)
        QtWidgets.QShortcut(QtGui.QKeySequence("-"), self).activated.connect(self.zoom_out)

    def zoom_in(self):
        if self.process.state() == QProcess.Running:
            self.process.write(b"+\n") 

    def zoom_out(self):
        if self.process.state() == QProcess.Running:
            self.process.write(b"-\n") 

    def get_val(self, widget):
        try: return widget.toPlainText()
        except: return widget.text()

    def init_hardware(self):
        try:
            p = os.path.join(self.BASE_DIR, "MobileNetSSD_deploy.prototxt")
            m = os.path.join(self.BASE_DIR, "MobileNetSSD_deploy.caffemodel")
            self.net = cv2.dnn.readNetFromCaffe(p, m)
        except: 
            self.net = None
        
        try:
            self.picam2 = Picamera2()
            config = self.picam2.create_preview_configuration(main={"size": (640, 480), "format": "RGB888"})
            self.picam2.configure(config)
            self.picam2.start()
        except: 
            print("Lỗi Camera Module 3")

        self.cam_worker = CameraWorker(self.picam2, getattr(self, 'net', None), self.CLASSES)
        self.cam_worker.frame_ready.connect(self.update_ui)
        self.cam_worker.start()

    def start_lidar(self):
        if self.process.state() == QProcess.Running: return
        os.system("sudo pkill -9 ultra_simple")
        prog = "/home/doan2026/Desktop/lidarC/rplidar_sdk/app/ultra_simple/ultra_simple"
        env = QtCore.QProcessEnvironment.systemEnvironment()
        env.insert("SDL_WINDOWID", str(int(self.lidarview.winId())))
        env.insert("SDL_VIDEODRIVER", "x11")
        self.process.setProcessEnvironment(env)
        self.process.start(prog)

    def stop_all(self):
        if self.process.state() == QProcess.Running: self.process.terminate()
        os.system("sudo pkill -9 ultra_simple")
        self.cam_scene.clear()
        self.khoangcach.setText(""); self.vatcan.setText("")

    def save_data_to_excel(self):
        try:
            now = QDateTime.currentDateTime().toString("HH:mm:ss dd-MM-yyyy")
            cam_p = os.path.join(self.BASE_DIR, "temp_cam.jpg")
            
            if self.latest_frame_save is not None:
                cv2.imwrite(cam_p, self.latest_frame_save)
            else:
                f = self.picam2.capture_array()
                if f is not None: cv2.imwrite(cam_p, f) 

            if not os.path.exists(self.EXCEL_FILE):
                wb = Workbook(); ws = wb.active
                ws.append(["Thời gian", "Vật cản", "Khoảng cách", "Hình ảnh"])
            else:
                wb = load_workbook(self.EXCEL_FILE); ws = wb.active

            row = ws.max_row + 1
            ws.cell(row=row, column=1).value = now
            ws.cell(row=row, column=2).value = self.get_val(self.vatcan)
            ws.cell(row=row, column=3).value = f"{self.get_val(self.khoangcach)}"
            
            if os.path.exists(cam_p):
                img = OpenpyxlImage(cam_p)
                img.width, img.height = 140, 105
                ws.add_image(img, f"D{row}")
            
            ws.row_dimensions[row].height = 85
            wb.save(self.EXCEL_FILE)
            QtWidgets.QMessageBox.information(self, "Lưu", f"Đã ghi bản ghi thành công lúc {now}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể lưu: {e}")

    def start_export_thread(self):
        if not os.path.exists(self.EXCEL_FILE):
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Hãy nhấn SAVE DATA trước!")
            return
        self.export_btn.setEnabled(False); self.export_btn.setText("Đang xuất...")
        self.thread = QThread()
        self.worker = ExportWorker(self.EXCEL_FILE)
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.on_export_finished)
        self.worker.finished.connect(self.thread.quit)
        self.thread.start()

    def on_export_finished(self, msg):
        self.export_btn.setEnabled(True); self.export_btn.setText("EXPORT")
        QtWidgets.QMessageBox.information(self, "Kết quả Export", msg)

    def read_lidar(self):
        while self.process.canReadLine():
            try:
                line = self.process.readLine().data().decode().strip()
                parts = line.split(',')
                if len(parts) == 360:
                    self.lidar_full_data = [float(x) for x in parts]
                    angles = list(range(355, 360)) + list(range(0, 6))
                    v_pts = [self.lidar_full_data[a] for a in angles if self.lidar_full_data[a] > 0]
                    if v_pts:
                        cur = int(min(v_pts))
                        if abs(cur - self.last_val) > 8:
                            self.khoangcach.setText(f"{cur} mm"); self.last_val = cur
                    else: self.khoangcach.setText("")
            except: pass

    def update_ui(self, frame_disp, found):
        try:
            self.latest_frame_save = frame_disp
            self.vatcan.setText(", ".join(list(set(found))) if found else "")
            
            h, w = frame_disp.shape[:2]
            rgb = np.ascontiguousarray(frame_disp)
            qt_img = QImage(rgb.data, w, h, w*3, QImage.Format_RGB888).rgbSwapped()
            self.cam_scene.clear()
            self.cam_scene.addPixmap(QPixmap.fromImage(qt_img))
        except: pass

    def closeEvent(self, event):
        try: self.cam_worker.stop()
        except: pass
        self.stop_all()
        event.accept()

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    win = LidarApp()
    win.show()
    sys.exit(app.exec_())
