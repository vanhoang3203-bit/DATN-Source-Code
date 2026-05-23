import sys, os
if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
else:
    base_dir = os.path.dirname(os.path.abspath(__file__))

log_path = os.path.join(base_dir, "nhat_ky_he_thong.txt")
sys.stdout = open(log_path, "w", encoding="utf-8", buffering=1)
sys.stderr = open(log_path, "w", encoding="utf-8", buffering=1)
# --------------------------------------------------

import socket, cv2, struct, shutil, numpy as np, threading, subprocess
from PyQt5 import QtWidgets, uic, QtCore, QtGui
from PyQt5.QtCore import QTimer, QDateTime, pyqtSignal, QObject, QThread, Qt
from PyQt5.QtGui import QImage, QPixmap, QPen, QColor, QFont
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from flask import Flask, render_template_string, Response, jsonify

flask_app = Flask(__name__)
win = None 

# Biến toàn cục lưu dữ liệu gửi về Web
lidar_current_data = []
current_obstacle = "CHỜ DỮ LIỆU"
current_distance = "-- mm"

# --- GIAO DIỆN WEB CHO ĐIỆN THOẠI ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>MÔ HÌNH XÁC ĐỊNH KHOẢNG CÁCH VÀ HÌNH DẠNG VẬT CẢN SỬ DỤNG LiDAR</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, sans-serif; background: #121212; color: white; margin: 0; padding: 10px; text-align: center;}
        h2 { margin: 10px 0 15px 0; color: #4facfe; font-size: 16px; text-transform: uppercase; line-height: 1.4;}
        
        /* Thiết kế Bảng thông số kỹ thuật Đồng bộ PC */
        .info-box { width: 100%; max-width: 450px; background: #1e1e1e; border: 2px solid #333; border-radius: 8px; padding: 12px; margin: 0 auto 15px auto; display: flex; justify-content: space-between; font-size: 14px; font-weight: bold; box-sizing: border-box;}
        .info-item { flex: 1; text-align: center; }
        .info-item span { display: block; font-size: 18px; margin-top: 5px; color: #4facfe; text-transform: uppercase;}

        .container { display: flex; flex-direction: row; flex-wrap: wrap; justify-content: center; width: 100%; gap: 15px; }
        .box { width: 100%; max-width: 450px; background: #000; border: 2px solid #333; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 8px rgba(0,0,0,0.5);}
        img { width: 100%; display: block; }
        canvas { display: block; width: 100%; max-width: 300px; margin: 0 auto;}
        
        .btn-group { display: flex; flex-wrap: wrap; gap: 10px; justify-content: center; width: 100%; max-width: 500px; margin: 20px auto; }
        button { flex: 1 1 40%; padding: 15px; border: none; border-radius: 8px; font-weight: bold; font-size: 14px; color: white; cursor: pointer; transition: 0.1s; }
        .btn-start { background: linear-gradient(45deg, #11998e, #38ef7d); }
        .btn-stop { background: linear-gradient(45deg, #cb2d3e, #ef473a); }
        .btn-zoom { background: linear-gradient(45deg, #4facfe, #00f2fe); }
        .btn-save { background: #4caf50; color: white; flex: 1 1 100%; }
        .btn-sync { background: #2196f3; color: white; flex: 1 1 100%; }
        button:active { transform: scale(0.95); opacity: 0.8; }
    </style>
</head>
<body>
    <h2>Mô Hình Xác Định Khoảng Cách<br>Và Hình Dạng Vật Cản Sử Dụng LiDAR</h2>
    
    <div class="info-box">
        <div class="info-item">VẬT CẢN<br><span id="val-vatcan">CHỜ DỮ LIỆU</span></div>
        <div class="info-item" style="border-left: 2px solid #333;">KHOẢNG CÁCH<br><span id="val-khoangcach">-- mm</span></div>
    </div>

    <div class="container">
        <div class="box"><img src="/video_feed" alt="Đang chờ Camera..."></div>
        <div class="box"><canvas id="lidarCanvas" width="300" height="300"></canvas></div>
    </div>
    
    <div class="btn-group">
        <button class="btn-start" onclick="sendCmd('start')">▶ BẮT ĐẦU</button>
        <button class="btn-stop" onclick="sendCmd('stop')">⏸ DỪNG</button>
        <button class="btn-zoom" onclick="sendCmd('zoomin')">🔍 ZOOM +</button>
        <button class="btn-zoom" onclick="sendCmd('zoomout')">🔍 ZOOM -</button>
        <button class="btn-save" onclick="sendCmd('save')">💾 1. LƯU DỮ LIỆU EXCEL</button>
        <button class="btn-sync" onclick="sendCmd('sync')">☁️ 2. XUẤT Ổ D & CLOUD</button>
    </div>

    <script>
        const canvas = document.getElementById('lidarCanvas');
        const ctx = canvas.getContext('2d');

        function drawLidar(points) {
            ctx.clearRect(0, 0, 300, 300);
            ctx.strokeStyle = '#004d00'; ctx.lineWidth = 1;
            for(let i=1; i<=5; i++) {
                ctx.beginPath(); ctx.arc(150, 150, i*30, 0, 2*Math.PI); ctx.stroke();
            }
            ctx.beginPath(); ctx.moveTo(150, 0); ctx.lineTo(150, 300); ctx.stroke();
            ctx.beginPath(); ctx.moveTo(0, 150); ctx.lineTo(300, 150); ctx.stroke();

            if(points && points.length > 0) {
                points.forEach((dist, angle) => {
                    if(dist > 150) {
                        let rad = (angle - 90) * Math.PI / 180;
                        let x = 150 + dist * 0.03 * Math.cos(rad); 
                        let y = 150 + dist * 0.03 * Math.sin(rad);
                        ctx.fillStyle = dist < 1000 ? 'red' : 'lime';
                        ctx.fillRect(x-1.5, y-1.5, 3, 3);
                    }
                });
            }
        }
        drawLidar([]);

        function updateMap() {
            fetch('/api/lidar')
                .then(res => res.json())
                .then(data => {
                    if(data.points && data.points.length > 0) { drawLidar(data.points); }
                    
                    const vVatCan = document.getElementById('val-vatcan');
                    const vKhoangCach = document.getElementById('val-khoangcach');
                    
                    vVatCan.innerText = data.obstacle;
                    vKhoangCach.innerText = data.distance;

                    if(data.obstacle === "CHƯA XÁC ĐỊNH") vVatCan.style.color = "#00FF00"; 
                    else if(data.obstacle !== "CHỜ DỮ LIỆU" && data.obstacle !== "TẠM DỪNG") vVatCan.style.color = "#FF3333"; 
                    else vVatCan.style.color = "#4facfe";

                    let distVal = parseInt(data.distance);
                    if(!isNaN(distVal)) {
                        vKhoangCach.style.color = distVal < 800 ? "#FF3333" : "#00FF00";
                    } else {
                        vKhoangCach.style.color = "#4facfe";
                    }
                }).catch(e => {});
        }
        setInterval(updateMap, 100);
        function sendCmd(cmd) { fetch('/api/cmd/' + cmd); }
    </script>
</body>
</html>
"""

# --- LUỒNG XỬ LÝ XUẤT DỮ LIỆU ---
class ExportWorker(QObject):
    finished = pyqtSignal(str)
    def __init__(self, excel_file):
        super().__init__()
        self.excel_file = excel_file

    def run(self):
        results = []
        try:
            cmd = f'rclone copy "{self.excel_file}" gdrive:DoAn_Lidar/'
            res = subprocess.run(cmd, shell=True, capture_output=True, text=True, encoding='utf-8')
            if res.returncode == 0:
                results.append("Cloud: OK")
            else:
                results.append(f"Cloud: Lỗi ({res.returncode})")
        except Exception as e:
            results.append("Cloud: Thất bại")

        try:
            d_path = r"D:\\"
            if os.path.exists(d_path):
                shutil.copy(self.excel_file, os.path.join(d_path, "BaoCao_KetQua_Lidar.xlsx"))
                results.append("Ổ D: OK")
            else:
                results.append("Ổ D: Trống")
        except:
            results.append("Ổ D: Lỗi")
        
        self.finished.emit(" | ".join(results))
# --- LUỒNG NHẬN VIDEO VÀ CHẠY AI ---
class CameraWorker(QThread):
    frame_ready = pyqtSignal(np.ndarray, list)
    def __init__(self, ip, net, ai_ready, app_instance):
        super().__init__()
        self.ip = ip
        self.net = net
        self.ai_ready = ai_ready
        self.app_instance = app_instance 
        self.running = True
        self.frame_count = 0
        self.last_detections = [] 
        self.CLASSES = ["background", "aeroplane", "bicycle", "bird", "boat", "bottle", "bus", "car", "cat", "chair", "cow", "diningtable", "dog", "horse", "motorbike", "person", "pottedplant", "sheep", "sofa", "train", "tvmonitor"]
    def resize_keep_aspect_fit(self, image, target_size):
        target_w, target_h = target_size
        h, w = image.shape[:2]
        scale = min(target_w/w, target_h/h)
        new_w, new_h = int(w * scale), int(h * scale)
        resized = cv2.resize(image, (new_w, new_h), interpolation=cv2.INTER_LANCZOS4)
        canvas = np.zeros((target_h, target_w, 3), dtype=np.uint8)
        x_offset = (target_w - new_w) // 2
        y_offset = (target_h - new_h) // 2
        canvas[y_offset:y_offset+new_h, x_offset:x_offset+new_w] = resized
        return canvas
    def run(self):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1.0) 
            sock.connect((self.ip, 5555))
            while self.running:
                try:
                    size_raw = sock.recv(struct.calcsize(">L"))
                    if not size_raw: break
                    size = struct.unpack(">L", size_raw)[0]
                    data = b""
                    while len(data) < size:
                        packet = sock.recv(size - len(data))
                        if not packet: break
                        data += packet
                except socket.timeout: continue 
                except Exception: break
                
                frame = cv2.imdecode(np.frombuffer(data, np.uint8), cv2.IMREAD_COLOR)
                if frame is None: continue
                frame = self.resize_keep_aspect_fit(frame, (731, 401))
                h, w = frame.shape[:2]
                found = []
                self.frame_count += 1
                if self.ai_ready and (self.frame_count % 2 == 0):
                    blob = cv2.dnn.blobFromImage(cv2.resize(frame, (300, 300)), 0.007843, (300, 300), 127.5)
                    self.net.setInput(blob)
                    detections = self.net.forward()
                    self.last_detections = []
                    for i in range(detections.shape[2]):
                        if detections[0, 0, i, 2] > 0.65:
                            idx = int(detections[0, 0, i, 1])
                            lbl = self.CLASSES[idx].upper()
                            box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
                            self.last_detections.append((lbl, box.astype("int")))
                for (lbl, (sX, sY, eX, eY)) in self.last_detections:
                    cv2.rectangle(frame, (sX, sY), (eX, eY), (0, 255, 0), 2)
                    cv2.putText(frame, lbl, (sX, sY-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                    found.append(lbl)
                self.frame_ready.emit(frame, found)
            sock.close()
        except: pass
    def stop(self): self.running = False
# --- GIAO DIỆN CHÍNH (PC) ---
class LidarWindowsApp(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.BASE_DIR = base_dir 
        uic.loadUi(os.path.join(self.BASE_DIR, "hienthi.ui"), self)
        
        self.PI_IP = '172.20.10.2'
        self.EXCEL_FILE = os.path.join(self.BASE_DIR, "data_log.xlsx")
        self.latest_frame_save = None
        self.last_val = 0
        self.l_sock = None
        self.lidar_buffer = ""
        self.zoom_factor = 1.0
        try:
            self.net = cv2.dnn.readNetFromCaffe(os.path.join(self.BASE_DIR, "MobileNetSSD_deploy.prototxt"), os.path.join(self.BASE_DIR, "MobileNetSSD_deploy.caffemodel"))
            self.ai_ready = True
        except: self.ai_ready = False
        self.cam_scene = QtWidgets.QGraphicsScene(self)
        self.CameraView.setScene(self.cam_scene)
        self.CameraView.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.CameraView.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.lidar_scene = QtWidgets.QGraphicsScene(self)
        self.lidarview.setScene(self.lidar_scene)
        self.lidarview.setBackgroundBrush(QtGui.QBrush(QColor(0, 0, 0)))
        self.lidarview.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.lidarview.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.t_l = QTimer(); self.t_l.timeout.connect(self.update_lidar)
        # CĂN GIỮA VÀ XÓA VIỀN NỀN CHO Ô VẬT CẢN / KHOẢNG CÁCH
        if hasattr(self, 'vatcan'):
            self.vatcan.setAlignment(Qt.AlignCenter)
            self.vatcan.setStyleSheet("background: transparent; border: none; color: white; font-weight: bold;")
        if hasattr(self, 'khoangcach'):
            self.khoangcach.setAlignment(Qt.AlignCenter)
            self.khoangcach.setStyleSheet("background: transparent; border: none; color: white; font-weight: bold;")

        # KẾT NỐI NÚT BẤM VỚI UI
        if hasattr(self, 'pushButton'): self.pushButton.clicked.connect(self.start_app)   
        if hasattr(self, 'pushButton_2'): self.pushButton_2.clicked.connect(self.stop_app) 
        if hasattr(self, 'save_btn'): self.save_btn.clicked.connect(self.save_data_to_excel)
        if hasattr(self, 'export_btn'): self.export_btn.clicked.connect(self.start_export_thread)

        if hasattr(self, 'zoomin'): self.zoomin.clicked.connect(self.zoom_in_btn)
        if hasattr(self, 'zoomout'): self.zoomout.clicked.connect(self.zoom_out_btn)
        
        if hasattr(self, 'verticalSlider'):
            self.verticalSlider.setMinimum(10)
            self.verticalSlider.setMaximum(30)
            self.verticalSlider.setValue(10)
            self.verticalSlider.valueChanged.connect(self.on_slider_zoom)
    def on_slider_zoom(self, value): 
        self.zoom_factor = value / 10.0
    @QtCore.pyqtSlot()
    def zoom_in_btn(self): 
        self.zoom_factor = min(3.0, self.zoom_factor + 0.2)
        if hasattr(self, 'verticalSlider'):
            self.verticalSlider.blockSignals(True)
            self.verticalSlider.setValue(int(self.zoom_factor * 10))
            self.verticalSlider.blockSignals(False)
    @QtCore.pyqtSlot()
    def zoom_out_btn(self): 
        self.zoom_factor = max(1.0, self.zoom_factor - 0.2)
        if hasattr(self, 'verticalSlider'):
            self.verticalSlider.blockSignals(True)
            self.verticalSlider.setValue(int(self.zoom_factor * 10))
            self.verticalSlider.blockSignals(False)
    def start_app(self):
        self.stop_app() 
        global current_obstacle, current_distance
        current_obstacle = "ĐANG QUÉT..."
        self.cam_worker = CameraWorker(self.PI_IP, getattr(self, 'net', None), self.ai_ready, self)
        self.cam_worker.frame_ready.connect(self.display_camera)
        self.cam_worker.start()
        try:
            self.l_sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.l_sock.connect((self.PI_IP, 5556))
            self.l_sock.setblocking(False)
            self.t_l.start(50)
        except: pass
    def stop_app(self):
        if hasattr(self, 'cam_worker'): self.cam_worker.stop()
        self.t_l.stop()
        if self.l_sock:
            try: self.l_sock.close()
            except: pass
            self.l_sock = None
        self.cam_scene.clear(); self.lidar_scene.clear()
        global current_obstacle, current_distance
        current_obstacle = "TẠM DỪNG"
        current_distance = "-- mm"
        if hasattr(self, 'khoangcach'): 
            self.khoangcach.setText("")
        if hasattr(self, 'vatcan'): 
            self.vatcan.setText("TẠM DỪNG")
            self.vatcan.setStyleSheet("background: transparent; border: none; color: white; font-weight: bold;")
    def display_camera(self, frame, found):
        h, w = frame.shape[:2]
        if self.zoom_factor > 1.0:
            new_h, new_w = int(h / self.zoom_factor), int(w / self.zoom_factor)
            top, left = (h - new_h) // 2, (w - new_w) // 2
            cropped = frame[top:top+new_h, left:left+new_w]
            frame_disp = cv2.resize(cropped, (w, h), interpolation=cv2.INTER_LANCZOS4).copy()
        else: frame_disp = frame.copy()
        frame_disp = cv2.convertScaleAbs(frame_disp, alpha=1.1, beta=5)
        gaussian = cv2.GaussianBlur(frame_disp, (0, 0), 2.0)
        frame_disp = cv2.addWeighted(frame_disp, 1.5, gaussian, -0.5, 0)
        self.latest_frame_save = frame_disp.copy()
        h_d, w_d = frame_disp.shape[:2]
        q_img = QImage(frame_disp.data, w_d, h_d, w_d*3, QImage.Format_RGB888).rgbSwapped()
        self.cam_scene.clear()
        self.cam_scene.addPixmap(QPixmap.fromImage(q_img))
        self.CameraView.fitInView(self.cam_scene.itemsBoundingRect(), Qt.KeepAspectRatio)
                txt = ", ".join(list(set(found))) if found else "CHƯA XÁC ĐỊNH"
        global current_obstacle
        current_obstacle = txt
        if hasattr(self, 'vatcan'):
            self.vatcan.setText(txt)
            if txt == "CHƯA XÁC ĐỊNH": self.vatcan.setStyleSheet("background: transparent; border: none; color: #00FF00; font-weight: bold;")
            else: self.vatcan.setStyleSheet("background: transparent; border: none; color: #FF3333; font-weight: bold;")
    def update_lidar(self):
        global lidar_current_data, current_distance
        try:
            raw = self.l_sock.recv(4096).decode()
            self.lidar_buffer += raw
            if '\n' in self.lidar_buffer:
                lines = self.lidar_buffer.split('\n'); self.lidar_buffer = lines[-1]
                parts = lines[-2].strip().split(',')
                if len(parts) == 360:
                    data = [float(x) for x in parts]
                    lidar_current_data = data
                    v_pts = [data[a] for a in (list(range(355, 360)) + list(range(0, 6))) if data[a] > 150]
                    if v_pts:
                        cur = int(min(v_pts))
                        current_distance = f"{cur} mm"
                        if abs(cur - self.last_val) > 8: 
                            if hasattr(self, 'khoangcach'):
                                self.khoangcach.setText(f"{cur} mm")
                                self.khoangcach.setStyleSheet("background: transparent; border: none; color: #FF3333; font-weight: bold;" if cur < 800 else "background: transparent; border: none; color: #00FF00; font-weight: bold;")
                            self.last_val = cur
                    self.draw_map(data)
        except: pass
    def draw_map(self, data):
        self.lidar_scene.clear(); self.lidar_scene.setSceneRect(-300, -300, 600, 600)
        p_grid = QPen(QColor(0, 80, 0)); f_font = QFont("Arial", 8, QFont.Bold)
        for r in range(1, 6):
            rad = r * 1000 * 0.05
            self.lidar_scene.addEllipse(-rad, -rad, rad*2, rad*2, p_grid)
            t = self.lidar_scene.addText(f"{r}m", f_font); t.setDefaultTextColor(Qt.white); t.setPos(rad, -15)
        self.lidar_scene.addLine(-300, 0, 300, 0, p_grid); self.lidar_scene.addLine(0, -300, 0, 300, p_grid)
        p_red = QPen(Qt.red, 3); p_green = QPen(Qt.green, 3)
        for a in range(360):
            dist = data[a]
            if dist > 150:
                rad_a = np.radians(a - 90)
                x = dist * 0.05 * np.cos(rad_a); y = dist * 0.05 * np.sin(rad_a)
                self.lidar_scene.addEllipse(x, y, 2, 2, p_red if dist < 1000 else p_green)
    def show_toast_message(self, title, message, timeout_ms=2000):
        self.toast = QtWidgets.QMessageBox(self)
        self.toast.setWindowTitle(title)
        self.toast.setText(message)
        self.toast.setTextFormat(Qt.PlainText)
        self.toast.setIcon(QtWidgets.QMessageBox.Information)
        self.toast.setWindowModality(Qt.NonModal)
        self.toast.setStyleSheet("""
            QMessageBox { background-color: #ffffff; border: 2px solid #2196F3; } 
            QLabel { color: #000000; background: transparent; font-size: 14px; font-weight: bold; } 
            QPushButton { background-color: #2196F3; color: white; padding: 6px 15px; font-weight: bold; border-radius: 4px; }
        """)
        self.toast.show()
        QtCore.QTimer.singleShot(timeout_ms, self.toast.accept)
    def save_data_to_excel(self):
        try:
            now = QDateTime.currentDateTime().toString("HH:mm:ss dd-MM-yyyy")
            img_p = os.path.join(self.BASE_DIR, "temp_cam.jpg")
            if self.latest_frame_save is not None: cv2.imwrite(img_p, self.latest_frame_save)
            wb = load_workbook(self.EXCEL_FILE) if os.path.exists(self.EXCEL_FILE) else Workbook()
            ws = wb.active
            if ws.max_row == 1 and ws['A1'].value is None: ws.append(["Thời gian", "Vật cản", "Khoảng cách", "Hình ảnh"])
            row = ws.max_row + 1
            vatcan_text = self.vatcan.text() if hasattr(self, 'vatcan') else ""
            khoangcach_text = self.khoangcach.text() if hasattr(self, 'khoangcach') else ""
            ws.cell(row=row, column=1, value=now)
            ws.cell(row=row, column=2, value=vatcan_text)
            ws.cell(row=row, column=3, value=khoangcach_text)
            if os.path.exists(img_p):
                img = OpenpyxlImage(img_p); img.width, img.height = 140, 105; ws.add_image(img, f"D{row}")
                ws.column_dimensions['D'].width = 20 
            ws.row_dimensions[row].height = 80
            wb.save(self.EXCEL_FILE)
            self.show_toast_message("Đã lưu", "File Excel đã được lưu trên máy tính!", 2000)
        except Exception as e: 
            self.show_toast_message("Lỗi", str(e), 3000)
    def start_export_thread(self):
        if not os.path.exists(self.EXCEL_FILE):
            self.show_toast_message("Cảnh báo", "Bạn chưa lưu file Excel!\nVui lòng bấm '1. LƯU DỮ LIỆU EXCEL' trước.", 3000)
            return
        if hasattr(self, 'export_btn'): self.export_btn.setEnabled(False)
        self.show_toast_message("Hệ thống", "Đang xử lý xuất dữ liệu ra Ổ D và Cloud...", 2000)
        self.export_thread = QThread()
        self.export_worker = ExportWorker(self.EXCEL_FILE)
        self.export_worker.moveToThread(self.export_thread)
        self.export_thread.started.connect(self.export_worker.run)
        self.export_worker.finished.connect(self.on_export_finished)
        self.export_worker.finished.connect(self.export_thread.quit)
        self.export_worker.finished.connect(self.export_worker.deleteLater)
        self.export_thread.finished.connect(self.export_thread.deleteLater)
        self.export_thread.start()
    def on_export_finished(self, message):
        self.show_toast_message("Báo cáo xuất", message, 3000)
        if hasattr(self, 'export_btn'): self.export_btn.setEnabled(True)
# --- WEB ROUTES ---
@flask_app.route('/')
def index(): return render_template_string(HTML_TEMPLATE)
def gen_frames():
    global win
    while True:
        if win is not None and win.latest_frame_save is not None:
            ret, buffer = cv2.imencode('.jpg', win.latest_frame_save)
            if ret: yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        QtCore.QThread.msleep(100) 
@flask_app.route('/video_feed')
def video_feed(): return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')
@flask_app.route('/api/lidar')
def api_lidar():
    global lidar_current_data, current_obstacle, current_distance
    return jsonify({
        'points': lidar_current_data,
        'obstacle': current_obstacle,
        'distance': current_distance
    })
@flask_app.route('/api/cmd/<command>')
def api_cmd(command):
    global win
    if win is None: return jsonify({"status": "error"})
    if command == 'start' and hasattr(win, 'pushButton'): QtCore.QMetaObject.invokeMethod(win.pushButton, "click", Qt.QueuedConnection)
    elif command == 'stop' and hasattr(win, 'pushButton_2'): QtCore.QMetaObject.invokeMethod(win.pushButton_2, "click", Qt.QueuedConnection)
    elif command == 'save' and hasattr(win, 'save_btn'): QtCore.QMetaObject.invokeMethod(win.save_btn, "click", Qt.QueuedConnection)
    elif command == 'sync' and hasattr(win, 'export_btn'): QtCore.QMetaObject.invokeMethod(win.export_btn, "click", Qt.QueuedConnection)
    elif command == 'zoomin': QtCore.QMetaObject.invokeMethod(win, "zoom_in_btn", Qt.QueuedConnection)
    elif command == 'zoomout': QtCore.QMetaObject.invokeMethod(win, "zoom_out_btn", Qt.QueuedConnection)
    return jsonify({"status": "ok"})
def run_flask():
    import logging; log = logging.getLogger('werkzeug'); log.setLevel(logging.ERROR)
    flask_app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    win = LidarWindowsApp()
    win.show() 
    threading.Thread(target=run_flask, daemon=True).start()
    sys.exit(app.exec_())
